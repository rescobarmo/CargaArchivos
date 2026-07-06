import re
from pathlib import Path
import pymysql
from pymysql.cursors import DictCursor
import logging

from config import MYSQL_CONFIG

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def _connect(use_db=True):
    cfg = {k: v for k, v in MYSQL_CONFIG.items()}
    if not use_db:
        cfg.pop("database", None)
    return pymysql.connect(**cfg, cursorclass=DictCursor)


def ensure_database():
    conn = _connect(use_db=False)
    try:
        db_name = MYSQL_CONFIG["database"]
        with conn.cursor() as cur:
            cur.execute(
                f"CREATE DATABASE IF NOT EXISTS `{db_name}` "
                f"CHARACTER SET utf8mb4 COLLATE utf8mb4_unicode_ci"
            )
        conn.commit()
    finally:
        conn.close()


def _sanitize_table_name(name: str) -> str:
    safe = re.sub(r"[^\w]", "_", name).strip("_").lower()
    if not safe or safe[0].isdigit():
        safe = "tbl_" + safe
    return safe[:60]


def _sanitize_column_name(name: str) -> str:
    safe = re.sub(r"[^\w]", "_", str(name)).strip("_").lower()
    if not safe:
        safe = "columna"
    if safe[0].isdigit():
        safe = "col_" + safe
    return safe[:60]


def _infer_mysql_type(value) -> str:
    if value is None:
        return None
    if isinstance(value, bool):
        return "TINYINT(1)"
    if isinstance(value, int):
        return "BIGINT"
    if isinstance(value, float):
        return "DOUBLE"
    
    s = str(value).strip()
    if not s:
        return None
    
    try:
        int(s)
        return "BIGINT"
    except (ValueError, TypeError):
        pass
    
    try:
        float(s)
        return "DOUBLE"
    except (ValueError, TypeError):
        pass
    
    if len(s) <= 255:
        return "VARCHAR(255)"
    return "TEXT"


def _infer_column_types(data_rows, num_columns):
    col_types = [None] * num_columns
    
    for row in data_rows:
        for i, val in enumerate(row):
            if i >= num_columns:
                break
            if val is None:
                continue
            
            inferred = _infer_mysql_type(val)
            
            if col_types[i] is None:
                col_types[i] = inferred
            elif inferred != col_types[i]:
                if col_types[i] in ("BIGINT", "DOUBLE") and inferred not in ("BIGINT", "DOUBLE"):
                    col_types[i] = "VARCHAR(255)"
                elif col_types[i] == "VARCHAR(255)" and inferred == "TEXT":
                    col_types[i] = "TEXT"
    
    for i in range(num_columns):
        if col_types[i] is None:
            col_types[i] = "VARCHAR(255)"
    
    return col_types


def import_excel_to_mysql(file_path, table_name: str) -> dict:
    ext = Path(file_path).suffix.lower()
    
    if ext == '.xls':
        return _import_xls(file_path, table_name)
    else:
        return _import_xlsx(file_path, table_name)


def _import_xls(file_path, table_name: str) -> dict:
    import xlrd
    
    logger.info(f"Importing XLS file: {file_path}")
    wb = xlrd.open_workbook(file_path)
    ws = wb.sheet_by_index(0)
    logger.info(f"Sheet has {ws.nrows} rows and {ws.ncols} columns")
    
    if ws.nrows == 0:
        return {"table": table_name, "rows": 0, "columns": 0, "error": "Archivo vacio"}
    
    raw_headers = [ws.cell_value(0, col) for col in range(ws.ncols)]
    logger.info(f"Raw headers: {raw_headers}")
    
    headers = []
    seen = {}
    for h in raw_headers:
        col = _sanitize_column_name(h if h else "columna")
        if col in seen:
            seen[col] += 1
            col = f"{col}_{seen[col]}"
        else:
            seen[col] = 0
        headers.append(col)
    logger.info(f"Sanitized headers: {headers}")
    
    if not headers:
        return {"table": table_name, "rows": 0, "columns": 0, "error": "Sin columnas"}
    
    data_rows = []
    for row_idx in range(1, ws.nrows):
        row = [ws.cell_value(row_idx, col) for col in range(ws.ncols)]
        data_rows.append(row)
    logger.info(f"Read {len(data_rows)} data rows")
    
    col_types = _infer_column_types(data_rows, len(headers))
    logger.info(f"Inferred column types: {col_types}")

    safe_table = _sanitize_table_name(table_name)
    logger.info(f"Table name: {safe_table}")
    
    ensure_database()
    conn = _connect()
    inserted = 0
    try:
        with conn.cursor() as cur:
            col_defs = ", ".join(
                f"`{h}` {col_types[i]}" for i, h in enumerate(headers)
            )
            cur.execute(f"DROP TABLE IF EXISTS `{safe_table}`")
            cur.execute(
                f"CREATE TABLE `{safe_table}` ({col_defs}) "
                f"ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci"
            )
            
            placeholders = ", ".join(["%s"] * len(headers))
            col_names = ", ".join(f"`{h}`" for h in headers)
            insert_sql = f"INSERT INTO `{safe_table}` ({col_names}) VALUES ({placeholders})"
            
            for row_idx, row in enumerate(data_rows):
                values = []
                for i, val in enumerate(row):
                    if i >= len(headers):
                        break
                    if isinstance(val, bool):
                        values.append(int(val))
                    elif val is None or val == "":
                        values.append(None)
                    else:
                        values.append(val)
                while len(values) < len(headers):
                    values.append(None)
                try:
                    cur.execute(insert_sql, values)
                    inserted += 1
                except Exception as e:
                    logger.error(f"Error inserting row {row_idx}: {e}, values: {values}")
                    raise

        conn.commit()
        logger.info(f"Successfully inserted {inserted} rows into {safe_table}")
    except Exception as e:
        logger.error(f"Database error: {e}")
        conn.rollback()
        raise
    finally:
        conn.close()
    
    return {
        "table": safe_table,
        "rows": inserted,
        "columns": len(headers),
    }


def _import_xlsx(file_path, table_name: str) -> dict:
    from openpyxl import load_workbook

    logger.info(f"Importing XLSX file: {file_path}")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    logger.info(f"Active sheet: {ws.title}")

    rows_iter = ws.iter_rows(values_only=True)

    raw_headers = next(rows_iter, None)
    if raw_headers is None:
        wb.close()
        return {"table": table_name, "rows": 0, "columns": 0, "error": "Archivo vacio"}
    
    logger.info(f"Raw headers: {raw_headers}")

    headers = []
    seen = {}
    for h in raw_headers:
        col = _sanitize_column_name(h if h is not None else "columna")
        if col in seen:
            seen[col] += 1
            col = f"{col}_{seen[col]}"
        else:
            seen[col] = 0
        headers.append(col)
    logger.info(f"Sanitized headers: {headers}")

    if not headers:
        wb.close()
        return {"table": table_name, "rows": 0, "columns": 0, "error": "Sin columnas"}

    data_rows = list(rows_iter)
    wb.close()
    logger.info(f"Read {len(data_rows)} data rows")

    col_types = _infer_column_types(data_rows, len(headers))
    logger.info(f"Inferred column types: {col_types}")

    safe_table = _sanitize_table_name(table_name)
    logger.info(f"Table name: {safe_table}")

    ensure_database()
    conn = _connect()
    inserted = 0
    try:
        with conn.cursor() as cur:
            col_defs = ", ".join(
                f"`{h}` {col_types[i]}" for i, h in enumerate(headers)
            )
            cur.execute(f"DROP TABLE IF EXISTS `{safe_table}`")
            cur.execute(
                f"CREATE TABLE `{safe_table}` ({col_defs}) "
                f"ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci"
            )

            placeholders = ", ".join(["%s"] * len(headers))
            col_names = ", ".join(f"`{h}`" for h in headers)
            insert_sql = f"INSERT INTO `{safe_table}` ({col_names}) VALUES ({placeholders})"

            for row_idx, row in enumerate(data_rows):
                values = []
                for i, val in enumerate(row):
                    if i >= len(headers):
                        break
                    if isinstance(val, bool):
                        values.append(int(val))
                    elif val is None or val == "":
                        values.append(None)
                    else:
                        values.append(val)
                while len(values) < len(headers):
                    values.append(None)
                try:
                    cur.execute(insert_sql, values)
                    inserted += 1
                except Exception as e:
                    logger.error(f"Error inserting row {row_idx}: {e}, values: {values}")
                    raise

        conn.commit()
        logger.info(f"Successfully inserted {inserted} rows into {safe_table}")
    except Exception as e:
        logger.error(f"Database error: {e}")
        conn.rollback()
        raise
    finally:
        conn.close()

    return {
        "table": safe_table,
        "rows": inserted,
        "columns": len(headers),
    }
